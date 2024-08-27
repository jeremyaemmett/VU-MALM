from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from warnings import filterwarnings
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import conversions
import openpyxl
import datetime
import params
import shutil
import init
import csv
import os

filterwarnings(action='ignore', category=DeprecationWarning, message='`np.bool` is a deprecated alias')

# test 8/21/2024

def reformat_date(dtime):

    print('dtime: ', dtime)
    year, day, month, hour, minute, second = \
        str(dtime.year), str(dtime.day), str(dtime.month), str(dtime.hour), str(dtime.minute), str(dtime.second)
    if len(day) == 1: day = '0' + day
    if len(month) == 1: month = '0' + month
    if len(hour) == 1: hour = '0' + hour
    if len(minute) == 1: minute = '0' + minute
    if len(second) == 1: second = '0' + second
    formatted_date = year + day + month + hour + minute + second

    return (formatted_date)


    if len(depths) > 0: # Depth profile
        with open(params.main_directory + 'siteData/standard/' + site + '/' + data + '/' + site + '_' + udate + extra_info + '.txt',
                "w") as current_file:
            current_file.write('Depths [cm], ' + data_info + '\n')
        with open(params.main_directory + 'siteData/standard/' + site + '/' + data + '/' + site + '_' + udate + extra_info + '.txt',
                "a") as current_file:
            for d in range(0, len(depths)):
                current_file.write(f"{depths[d]},{values[d]}\n")
    else: # Single value
        with open(params.main_directory + 'siteData/standard/' + site + '/' + data + '/' + site + '_' + udate + extra_info + '.txt',
                "w") as current_file:
            current_file.write(data_info + '\n')
        with open(params.main_directory + 'siteData/standard/' + site + '/' + data + '/' + site + '_' + udate + extra_info + '.txt',
                "a") as current_file:
            current_file.write(str(values[0]))


    return()


def write_standard2(site, std_folder_name, date_string, depths, values, std_depth_header, std_value_header, extra_info):

    print(depths, values)
    single_value_check = len(depths) == 1

    if len(depths) > 0: # Depth profile

        with open(params.main_directory + 'siteData/standard/' + site + '/' + std_folder_name + '/' + site + '_' + date_string + extra_info + '.txt',
                "w") as current_file:

            if single_value_check == False:
                current_file.write(std_depth_header + ', ' + std_value_header + '\n')
            else:
                current_file.write(std_depth_header + '\n')

        with open(params.main_directory + 'siteData/standard/' + site + '/' + std_folder_name + '/' + site + '_' + date_string + extra_info + '.txt',
                "a") as current_file:
            if single_value_check == False:
                for d in range(0, len(depths)):
                    current_file.write(f"{depths[d]},{values[d]}\n")
            else:
                current_file.write(f"{values[0]}\n")

    return()


def standardization_method_1(site, std_folder_name, date_string, raw_filename, raw_worksheet,
    raw_depth_column, raw_value_column, raw_depth_row1, raw_depth_row2, raw_value_row1, raw_value_row2,
    std_depth_header, std_value_header, raw2std_depth_cf, raw2std_value_cf):

    path = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/standard/'+site+'/'
    if not os.path.exists(path): os.mkdir(path)
    path = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/standard/'+site+'/'+std_folder_name+'/'
    if not os.path.exists(path): os.mkdir(path)

    # Get the column index from the Excel-style column string
    value_col_idx = column_index_from_string(coordinate_from_string(raw_value_column + '1')[0])
    depth_col_idx = column_index_from_string(coordinate_from_string(raw_depth_column + '1')[0])

    wb = load_workbook(filename=raw_filename, data_only = True)
    ws = wb[raw_worksheet]

    # List data as it appears in the file, as string elements
    values_as_strings = [ws.cell(row=i, column=value_col_idx).value for i in range(raw_value_row1, raw_value_row2)]
    depths_as_strings = [ws.cell(row=i, column=depth_col_idx).value for i in range(raw_depth_row1, raw_depth_row2)]

    # Check the data type of the independent variable, for handling
    independent_data_type = type(depths_as_strings[0])

    # Convert lists to arrays, as float elements, excluding data gaps
    print('1: ', depths_as_strings)
    print(values_as_strings)
    values = np.asarray([float(value) for value in values_as_strings if type(value) != str])
    if independent_data_type != datetime.datetime:
        depths = np.asarray([float(depths_as_strings[i]) for i in range(0, len(depths_as_strings)) if type(values_as_strings[i]) != str])
        depths = raw2std_depth_cf * depths
    else:
        # Reformat the datetimes to a strings with format: YYYYDDMMHHMMSS
        print(site, date_string)
        depths_as_strings = [reformat_date(d) for d in depths_as_strings]
        depths = np.asarray(depths_as_strings)

    values = raw2std_value_cf * values

    print(depths)
    print(values)

    # Make the output folder for the standardized data, if necessary
    path = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/standard/'+site+'/'+std_folder_name
    if not os.path.exists(path): os.mkdir(path)

    print('std_folder_name: ', std_folder_name)
    write_standard2(site, std_folder_name, date_string, depths, values, std_depth_header, std_value_header, '')

    return()


def standardization_method_2(site, std_folder_name, date_string, raw_filename, raw_worksheet,
    raw_plot_column, raw_depth_column, raw_value_column,
    raw_plot_row1, raw_plot_row2, raw_depth_row1, raw_depth_row2, raw_value_row1, raw_value_row2,
    std_depth_header, std_value_header, raw2std_depth_cf, raw2std_value_cf):

    path = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/standard/'+site+'/'
    if not os.path.exists(path): os.mkdir(path)
    path = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/standard/'+site+'/'+std_folder_name+'/'
    if not os.path.exists(path): os.mkdir(path)

    # Get the column index from the Excel-style column string
    plot_col_idx = column_index_from_string(coordinate_from_string(raw_plot_column + '1')[0])
    value_col_idx = column_index_from_string(coordinate_from_string(raw_value_column + '1')[0])
    depth_col_idx = column_index_from_string(coordinate_from_string(raw_depth_column + '1')[0])

    wb = load_workbook(filename=raw_filename, data_only = True)
    ws = wb[raw_worksheet]

    # List data as it appears in the file, as string elements
    plots_as_strings = [ws.cell(row=i, column=plot_col_idx).value for i in range(raw_plot_row1, raw_plot_row2)]
    values_as_strings = [ws.cell(row=i, column=value_col_idx).value for i in range(raw_value_row1, raw_value_row2)]
    depths_as_strings = [ws.cell(row=i, column=depth_col_idx).value for i in range(raw_depth_row1, raw_depth_row2)]

    day, month, year = int(date_string[0:2]), int(date_string[2:4]), int(date_string[4:8])

    d1 = datetime.datetime(year, month, day)
    d2 = d1 + datetime.timedelta(days = 14)

    # Select 'Light' (L) chamber measurements, taken from plots in the 'site' group, and < 2 weeks after the start date
    depths = np.asarray([depths_as_strings[p] for p in range(0, len(depths_as_strings))
                    if plots_as_strings[p][-1] == 'L' and plots_as_strings[p][:-1] in params.subplots[site + '_plots']
                    and d1 < depths_as_strings[p] < d2])
    values = np.asarray([values_as_strings[p] for p in range(0, len(values_as_strings))
                    if plots_as_strings[p][-1] == 'L' and plots_as_strings[p][:-1] in params.subplots[site + '_plots']
                    and d1 < depths_as_strings[p] < d2])

    # Sort dates and fluxes, by date
    values = [x for _, x in sorted(zip(depths, values))]
    depths = sorted(depths)

    # Reformat the datetimes to a strings with format: YYYYDDMMHHMMSS
    depths = [reformat_date(d) for d in depths]

    # Remove outliers using a 1-sigma threshold
    depths = [depths[x] for x in range(0, len(depths)) if (values[x] > np.mean(values) - np.std(values)) and (values[x] < np.mean(values) + np.std(values))]
    values = [x for x in values if (x > np.mean(values) - np.std(values)) and (x < np.mean(values) + np.std(values))]

    print(site, date_string, ' mean: ', np.mean(values))
    print(site, date_string, ' stdv: ', np.std(values)/4.0)

    write_standard2(site, std_folder_name, date_string, depths, values, std_depth_header, std_value_header, '')

    return()


def standardization_method_3(site, std_folder_name, date_string, raw_filename, raw_worksheet,
    raw_depth_column, raw_value_column, raw_depth_row1, raw_depth_row2, raw_value_row1, raw_value_row2,
    std_depth_header, std_value_header, raw2std_depth_cf, raw2std_value_cf):

    path = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/standard/'+site+'/'
    if not os.path.exists(path): os.mkdir(path)
    path = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/standard/'+site+'/'+std_folder_name+'/'
    if not os.path.exists(path): os.mkdir(path)

    # Get the column index from the Excel-style column string
    value_col_idx = column_index_from_string(coordinate_from_string(raw_value_column + '1')[0])
    depth_col_idx = column_index_from_string(coordinate_from_string(raw_depth_column + '1')[0])

    wb = load_workbook(filename=raw_filename, data_only = True)
    ws = wb[raw_worksheet]

    # List data as it appears in the file, as string elements
    values_as_strings = [ws.cell(row=i, column=value_col_idx).value for i in range(raw_value_row1, raw_value_row2)]
    depths_as_strings = [ws.cell(row=i, column=depth_col_idx).value for i in range(raw_depth_row1, raw_depth_row2)]

    # Check the data type of the independent variable, for handling
    independent_data_type = type(depths_as_strings[0])

    # Convert lists to arrays, as float elements, excluding data gaps
    values = np.asarray([float(value) for value in values_as_strings if type(value) != str])
    if independent_data_type != datetime.datetime:
        depths = np.asarray([float(depths_as_strings[i]) for i in range(0, len(depths_as_strings)) if type(values_as_strings[i]) != str])
        depths = raw2std_depth_cf * depths
    else:
        depths = np.asarray(depths_as_strings)

    values = raw2std_value_cf * values

    # Adjust the y-values to follow an exponential function
    log_x_data = np.log(depths)
    curve = np.polyfit(log_x_data, values, 1)
    curve_values = curve[0] * np.log(depths) + curve[1]
    curve_values[curve_values < 0.0] = 0.0

    print(depths)
    print(curve_values)

    # Make the output folder for the standardized data, if necessary
    path = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/standard/'+site+'/'+std_folder_name
    if not os.path.exists(path): os.mkdir(path)

    print('std_folder_name: ', std_folder_name)
    write_standard2(site, std_folder_name, date_string, depths, curve_values, std_depth_header, std_value_header, '')
    # 8/19/2024
    return()


def read_site_data():

    composition_file, composition_worksheet = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/raw/composition_raw.xlsx', 'composition_raw'
    genes_file, genes_worksheet = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/raw/genes_raw.xlsx', 'Sheet1'
    temps_file, temps_worksheet = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/raw/temps_raw.xlsx', 'Sheet1'
    fluxes_file, fluxes_worksheet = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/raw/fluxes_raw.xlsx', 'sorted_fluxes'
    roots_file, roots_worksheet = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/raw/roots_raw.xlsx', 'Sheet1'
    geography_file, geography_worksheet = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/raw/geography_raw.xlsx', 'Sheet1'

    print('Standardizing raw site data: ', 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/raw/<site>/<data>.csv')

    #if params.std_roots: read_roots(site, 'roots', 'Root Mass Density [kg/m^3]')

    # Input format:
    # [site name], [std data directory name], [date of profile (ddmmyyyy)],
    # [raw data file path],
    # [raw data depth label], [raw data value label], [raw data depth column], [raw data value column],
    # [raw data depth row1], [raw data depth row2], [raw data value row1], [raw data value row 2],
    # [std data depth label], [std data value label],
    # [raw2std depth con.f.], [raw2std value con.f.]

    # Geography
    if params.std_geography:
        #
        standardization_method_1('palsa_low', 'albedo', '15082023',
            geography_file, geography_worksheet, 'E', 'E', 2, 3, 2, 3, 'Albedo', 'Albedo', 1.0, 1.0)
        standardization_method_1('palsa_low', 'elevation', '15082023',
            geography_file, geography_worksheet, 'D', 'D', 2, 3, 2, 3, 'Elevation (m)', 'Elevation (m)', 1.0, 1.0)
        standardization_method_1('palsa_low', 'latitude', '15082023',
            geography_file, geography_worksheet, 'B', 'B', 2, 3, 2, 3, 'Latitude (degN)', 'Latitude (degN)', 1.0, 1.0)
        standardization_method_1('palsa_low', 'longitude', '15082023',
            geography_file, geography_worksheet, 'C', 'C', 2, 3, 2, 3, 'Longitude (degE)', 'Longitude (degE)', 1.0, 1.0)
        #
        standardization_method_1('palsa_high', 'albedo', '15082023',
            geography_file, geography_worksheet, 'E', 'E', 3, 4, 3, 4, 'Albedo', 'Albedo', 1.0, 1.0)
        standardization_method_1('palsa_high', 'elevation', '15082023',
            geography_file, geography_worksheet, 'D', 'D', 3, 4, 3, 4, 'Elevation (m)', 'Elevation (m)', 1.0, 1.0)
        standardization_method_1('palsa_high', 'latitude', '15082023',
            geography_file, geography_worksheet, 'B', 'B', 3, 4, 3, 4, 'Latitude (degN)', 'Latitude (degN)', 1.0, 1.0)
        standardization_method_1('palsa_high', 'longitude', '15082023',
            geography_file, geography_worksheet, 'C', 'C', 3, 4, 3, 4, 'Longitude (degE)', 'Longitude (degE)', 1.0, 1.0)
        #
        standardization_method_1('fen_low', 'albedo', '15082023',
            geography_file, geography_worksheet, 'E', 'E', 4, 5, 4, 5, 'Albedo', 'Albedo', 1.0, 1.0)
        standardization_method_1('fen_low', 'elevation', '15082023',
            geography_file, geography_worksheet, 'D', 'D', 4, 5, 4, 5, 'Elevation (m)', 'Elevation (m)', 1.0, 1.0)
        standardization_method_1('fen_low', 'latitude', '15082023',
            geography_file, geography_worksheet, 'B', 'B', 4, 5, 4, 5, 'Latitude (degN)', 'Latitude (degN)', 1.0, 1.0)
        standardization_method_1('fen_low', 'longitude', '15082023',
            geography_file, geography_worksheet, 'C', 'C', 4, 5, 4, 5, 'Longitude (degE)', 'Longitude (degE)', 1.0, 1.0)
        #
        standardization_method_1('fen_high', 'albedo', '15082023',
            geography_file, geography_worksheet, 'E', 'E', 5, 6, 5, 6, 'Albedo', 'Albedo', 1.0, 1.0)
        standardization_method_1('fen_high', 'elevation', '15082023',
            geography_file, geography_worksheet, 'D', 'D', 5, 6, 5, 6, 'Elevation (m)', 'Elevation (m)', 1.0, 1.0)
        standardization_method_1('fen_high', 'latitude', '15082023',
            geography_file, geography_worksheet, 'B', 'B', 5, 6, 5, 6, 'Latitude (degN)', 'Latitude (degN)', 1.0, 1.0)
        standardization_method_1('fen_high', 'longitude', '15082023',
            geography_file, geography_worksheet, 'C', 'C', 5, 6, 5, 6, 'Longitude (degE)', 'Longitude (degE)', 1.0, 1.0)
        #
        standardization_method_1('tundra_low', 'albedo', '15082023',
            geography_file, geography_worksheet, 'E', 'E', 6, 7, 6, 7, 'Albedo', 'Albedo', 1.0, 1.0)
        standardization_method_1('tundra_low', 'elevation', '15082023',
            geography_file, geography_worksheet, 'D', 'D', 6, 7, 6, 7, 'Elevation (m)', 'Elevation (m)', 1.0, 1.0)
        standardization_method_1('tundra_low', 'latitude', '15082023',
            geography_file, geography_worksheet, 'B', 'B', 6, 7, 6, 7, 'Latitude (degN)', 'Latitude (degN)', 1.0, 1.0)
        standardization_method_1('tundra_low', 'longitude', '15082023',
            geography_file, geography_worksheet, 'C', 'C', 6, 7, 6, 7, 'Longitude (degE)', 'Longitude (degE)', 1.0, 1.0)
        #
        standardization_method_1('tundra_high', 'albedo', '15082023',
            geography_file, geography_worksheet, 'E', 'E', 7, 8, 7, 8, 'Albedo', 'Albedo', 1.0, 1.0)
        standardization_method_1('tundra_high', 'elevation', '15082023',
            geography_file, geography_worksheet, 'D', 'D', 7, 8, 7, 8, 'Elevation (m)', 'Elevation (m)', 1.0, 1.0)
        standardization_method_1('tundra_high', 'latitude', '15082023',
            geography_file, geography_worksheet, 'B', 'B', 7, 8, 7, 8, 'Latitude (degN)', 'Latitude (degN)', 1.0, 1.0)
        standardization_method_1('tundra_high', 'longitude', '15082023',
            geography_file, geography_worksheet, 'C', 'C', 7, 8, 7, 8, 'Longitude (degE)', 'Longitude (degE)', 1.0, 1.0)

    # VWR
    if params.std_waterpct:
        standardization_method_1('palsa_low', 'waterpct', '15082023',
            composition_file, composition_worksheet, 'AO', 'AP', 15, 24, 15, 24, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('palsa_high', 'waterpct', '15082023',
            composition_file, composition_worksheet, 'AO', 'AQ', 15, 24, 15, 24, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('fen_high', 'waterpct', '15082023',
            composition_file, composition_worksheet, 'AO', 'AP', 3, 12, 3, 12, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('fen_low', 'waterpct', '15082023',
            composition_file, composition_worksheet, 'AO', 'AQ', 3, 12, 3, 12, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('tundra_high', 'waterpct', '15082023',
            composition_file, composition_worksheet, 'AO', 'AP', 27, 36, 27, 36, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('tundra_low', 'waterpct', '15082023',
            composition_file, composition_worksheet, 'AO', 'AQ', 27, 36, 27, 36, 'Depth (m)', 'Water (%)', 0.01, 0.01)

    # MIN
    if params.std_mineralpct:
        standardization_method_1('palsa_low', 'mineralpct', '15082023',
            composition_file, composition_worksheet, 'Z', 'AA', 15, 24, 15, 24, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('palsa_high', 'mineralpct', '15082023',
            composition_file, composition_worksheet, 'Z', 'AB', 15, 24, 15, 24, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('fen_high', 'mineralpct', '15082023',
            composition_file, composition_worksheet, 'Z', 'AA', 3, 12, 3, 12, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('fen_low', 'mineralpct', '15082023',
            composition_file, composition_worksheet, 'Z', 'AB', 3, 12, 3, 12, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('tundra_high', 'mineralpct', '15082023',
            composition_file, composition_worksheet, 'Z', 'AA', 27, 36, 27, 36, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('tundra_low', 'mineralpct', '15082023',
            composition_file, composition_worksheet, 'Z', 'AB', 27, 36, 27, 36, 'Depth (m)', 'Water (%)', 0.01, 0.01)

    # SOC
    if params.std_organicpct:
        standardization_method_1('palsa_low', 'organicpct', '15082023',
            composition_file, composition_worksheet, 'AC', 'AD', 15, 24, 15, 24, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('palsa_high', 'organicpct', '15082023',
            composition_file, composition_worksheet, 'AC', 'AE', 15, 24, 15, 24, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('fen_high', 'organicpct', '15082023',
            composition_file, composition_worksheet, 'AC', 'AD', 3, 12, 3, 12, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('fen_low', 'organicpct', '15082023',
            composition_file, composition_worksheet, 'AC', 'AE', 3, 12, 3, 12, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('tundra_high', 'organicpct', '15082023',
            composition_file, composition_worksheet, 'AC', 'AD', 27, 36, 27, 36, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('tundra_low', 'organicpct', '15082023',
            composition_file, composition_worksheet, 'AC', 'AE', 27, 36, 27, 36, 'Depth (m)', 'Water (%)', 0.01, 0.01)

    # AIR
    if params.std_airpct:
        standardization_method_1('palsa_low', 'airpct', '15082023',
            composition_file, composition_worksheet, 'AR', 'AS', 15, 24, 15, 24, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('palsa_high', 'airpct', '15082023',
            composition_file, composition_worksheet, 'AR', 'AT', 15, 24, 15, 24, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('fen_high', 'airpct', '15082023',
            composition_file, composition_worksheet, 'AR', 'AS', 3, 12, 3, 12, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('fen_low', 'airpct', '15082023',
            composition_file, composition_worksheet, 'AR', 'AT', 3, 12, 3, 12, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('tundra_high', 'airpct', '15082023',
            composition_file, composition_worksheet, 'AR', 'AS', 27, 36, 27, 36, 'Depth (m)', 'Water (%)', 0.01, 0.01)
        standardization_method_1('tundra_low', 'airpct', '15082023',
            composition_file, composition_worksheet, 'AR', 'AT', 27, 36, 27, 36, 'Depth (m)', 'Water (%)', 0.01, 0.01)

    # Water Table
    if params.std_watertable:
        # August
        standardization_method_1('palsa_low', 'watertable', '15082023',
            composition_file, composition_worksheet, 'AL', 'AM', 15, 24, 15, 24, 'Depth (m)', 'Water (%)', 0.01, 1.0)
        standardization_method_1('palsa_high', 'watertable', '15082023',
            composition_file, composition_worksheet, 'AL', 'AN', 15, 24, 15, 24, 'Depth (m)', 'Water (%)', 0.01, 1.0)
        standardization_method_1('fen_high', 'watertable', '15082023',
            composition_file, composition_worksheet, 'AL', 'AM', 3, 12, 3, 12, 'Depth (m)', 'Water (%)', 0.01, 1.0)
        standardization_method_1('fen_low', 'watertable', '15082023',
            composition_file, composition_worksheet, 'AL', 'AN', 3, 12, 3, 12, 'Depth (m)', 'Water (%)', 0.01, 1.0)
        standardization_method_1('tundra_high', 'watertable', '15082023',
            composition_file, composition_worksheet, 'AL', 'AM', 27, 36, 27, 36, 'Depth (m)', 'Water (%)', 0.01, 1.0)
        standardization_method_1('tundra_low', 'watertable', '15082023',
            composition_file, composition_worksheet, 'AL', 'AN', 27, 36, 27, 36, 'Depth (m)', 'Water (%)', 0.01, 1.0)
        # June
        standardization_method_1('palsa_low', 'watertable', '23062023',
            composition_file, composition_worksheet, 'AL', 'AM', 51, 60, 51, 60, 'Depth (m)', 'Water (%)', 0.01, 1.0)
        standardization_method_1('palsa_high', 'watertable', '23062023',
            composition_file, composition_worksheet, 'AL', 'AN', 51, 60, 51, 60, 'Depth (m)', 'Water (%)', 0.01, 1.0)
        standardization_method_1('fen_high', 'watertable', '23062023',
            composition_file, composition_worksheet, 'AL', 'AM', 39, 48, 39, 48, 'Depth (m)', 'Water (%)', 0.01, 1.0)
        standardization_method_1('fen_low', 'watertable', '23062023',
            composition_file, composition_worksheet, 'AL', 'AN', 39, 48, 39, 48, 'Depth (m)', 'Water (%)', 0.01, 1.0)
        standardization_method_1('tundra_high', 'watertable', '23062023',
            composition_file, composition_worksheet, 'AL', 'AM', 63, 72, 63, 72, 'Depth (m)', 'Water (%)', 0.01, 1.0)
        standardization_method_1('tundra_low', 'watertable', '23062023',
            composition_file, composition_worksheet, 'AL', 'AN', 63, 72, 63, 72, 'Depth (m)', 'Water (%)', 0.01, 1.0)

    # mcrA
    if params.std_mcra:
            # June
        standardization_method_1('palsa_low', 'mcra', '15062023',
            genes_file, genes_worksheet, 'AE', 'AF', 23, 39, 23, 39, 'Depth (m)', 'mcrA (copies/dry g)', 0.01, 1.0)
        standardization_method_1('palsa_high', 'mcra', '15062023',
            genes_file, genes_worksheet, 'AE', 'AG', 23, 39, 23, 39, 'Depth (m)', 'mcrA (copies/dry g)', 0.01, 1.0)
        standardization_method_1('fen_low', 'mcra', '15062023',
            genes_file, genes_worksheet, 'AE', 'AF', 3, 20, 3, 20, 'Depth (m)', 'mcrA (copies/dry g)', 0.01, 1.0)
        standardization_method_1('fen_high', 'mcra', '15062023',
            genes_file, genes_worksheet, 'AE', 'AG', 3, 20, 3, 20, 'Depth (m)', 'mcrA (copies/dry g)', 0.01, 1.0)
        standardization_method_1('tundra_low', 'mcra', '15062023',
            genes_file, genes_worksheet, 'AE', 'AF', 43, 60, 43, 60, 'Depth (m)', 'mcrA (copies/dry g)', 0.01, 1.0)
        standardization_method_1('tundra_high', 'mcra', '15062023',
            genes_file, genes_worksheet, 'AE', 'AG', 43, 60, 43, 60, 'Depth (m)', 'mcrA (copies/dry g)', 0.01, 1.0)
            # August
        standardization_method_1('palsa_low', 'mcra', '15082023',
            genes_file, genes_worksheet, 'AE', 'AF', 83, 100, 83, 100, 'Depth (m)', 'mcrA (copies/dry g)', 0.01, 1.0)
        standardization_method_1('palsa_high', 'mcra', '15082023',
            genes_file, genes_worksheet, 'AE', 'AG', 83, 100, 83, 100, 'Depth (m)', 'mcrA (copies/dry g)', 0.01, 1.0)
        standardization_method_1('fen_low', 'mcra', '15082023',
            genes_file, genes_worksheet, 'AE', 'AF', 63, 80, 63, 80, 'Depth (m)', 'mcrA (copies/dry g)', 0.01, 1.0)
        standardization_method_1('fen_high', 'mcra', '15082023',
            genes_file, genes_worksheet, 'AE', 'AG', 63, 80, 63, 80, 'Depth (m)', 'mcrA (copies/dry g)', 0.01, 1.0)
        standardization_method_1('tundra_low', 'mcra', '15082023',
            genes_file, genes_worksheet, 'AE', 'AF', 103, 120, 103, 120, 'Depth (m)', 'mcrA (copies/dry g)', 0.01, 1.0)
        standardization_method_1('tundra_high', 'mcra', '15082023',
            genes_file, genes_worksheet, 'AE', 'AG', 103, 120, 103, 120, 'Depth (m)', 'mcrA (copies/dry g)', 0.01, 1.0)

    # pmoA2
    if params.std_pmoa2:
            # June
        standardization_method_1('palsa_low', 'pmoa2', '15062023',
            genes_file, genes_worksheet, 'AH', 'AI', 23, 39, 23, 39, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('palsa_high', 'pmoa2', '15062023',
            genes_file, genes_worksheet, 'AH', 'AJ', 23, 39, 23, 39, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('fen_low', 'pmoa2', '15062023',
            genes_file, genes_worksheet, 'AH', 'AI', 3, 20, 3, 20, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('fen_high', 'pmoa2', '15062023',
            genes_file, genes_worksheet, 'AH', 'AJ', 3, 20, 3, 20, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('tundra_low', 'pmoa2', '15062023',
            genes_file, genes_worksheet, 'AH', 'AI', 43, 60, 43, 60, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('tundra_high', 'pmoa2', '15062023',
            genes_file, genes_worksheet, 'AH', 'AJ', 43, 60, 43, 60, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
            # August
        standardization_method_1('palsa_low', 'pmoa2', '15082023',
            genes_file, genes_worksheet, 'AH', 'AI', 83, 100, 83, 100, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('palsa_high', 'pmoa2', '15082023',
            genes_file, genes_worksheet, 'AH', 'AJ', 83, 100, 83, 100, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('fen_low', 'pmoa2', '15082023',
            genes_file, genes_worksheet, 'AH', 'AI', 63, 80, 63, 80, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('fen_high', 'pmoa2', '15082023',
            genes_file, genes_worksheet, 'AH', 'AJ', 63, 80, 63, 80, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('tundra_low', 'pmoa2', '15082023',
            genes_file, genes_worksheet, 'AH', 'AI', 103, 120, 103, 120, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('tundra_high', 'pmoa2', '15082023',
            genes_file, genes_worksheet, 'AH', 'AJ', 103, 120, 103, 120, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)

    # pmoA1
    if params.std_pmoa1:
            # June
        standardization_method_1('palsa_low', 'pmoa1', '15062023',
            genes_file, genes_worksheet, 'AK', 'AL', 23, 39, 23, 39, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('palsa_high', 'pmoa1', '15062023',
            genes_file, genes_worksheet, 'AK', 'AM', 23, 39, 23, 39, 'Depth (m)',  'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('fen_low', 'pmoa1', '15062023',
            genes_file, genes_worksheet, 'AK', 'AL', 3, 20, 3, 20, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('fen_high', 'pmoa1', '15062023',
            genes_file, genes_worksheet, 'AK', 'AM', 3, 20, 3, 20, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('tundra_low', 'pmoa1', '15062023',
            genes_file, genes_worksheet, 'AK', 'AL', 43, 60, 43, 60, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('tundra_high', 'pmoa1', '15062023',
            genes_file, genes_worksheet, 'AK', 'AM', 43, 60, 43, 60, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
            # August
        standardization_method_1('palsa_low', 'pmoa1', '15082023',
            genes_file, genes_worksheet, 'AK', 'AL', 83, 100, 83, 100, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('palsa_high', 'pmoa1', '15082023',
            genes_file, genes_worksheet, 'AK', 'AM', 83, 100, 83, 100, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('fen_low', 'pmoa1', '15082023',
            genes_file, genes_worksheet, 'AK', 'AL', 63, 80, 63, 80, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('fen_high', 'pmoa1', '15082023',
            genes_file, genes_worksheet, 'AK', 'AM', 63, 80, 63, 80, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('tundra_low', 'pmoa1', '15082023',
            genes_file, genes_worksheet, 'AK', 'AL', 103, 120, 103, 120, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)
        standardization_method_1('tundra_high', 'pmoa1', '15082023',
            genes_file, genes_worksheet, 'AK', 'AM', 103, 120, 103, 120, 'Depth (m)', 'pmoA2 (copies/dry g)', 0.01, 1.0)

    # Temps
    if params.std_temps:
            # 5 cm - test
        standardization_method_1('palsa_low', 'temps', '15062023_05cm',
            temps_file, temps_worksheet, 'A', 'K', 3, 505, 3, 505, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('palsa_high', 'temps', '15062023_05cm',
            temps_file, temps_worksheet, 'A', 'AS', 3, 505, 3, 505, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('fen_low', 'temps', '15062023_05cm',
            temps_file, temps_worksheet, 'AY', 'BI', 3, 444, 3, 444, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('fen_high', 'temps', '15062023_05cm',
            temps_file, temps_worksheet, 'AY', 'CQ', 3, 444, 3, 444, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('tundra_low', 'temps', '15062023_05cm',
            temps_file, temps_worksheet, 'CW', 'DG', 3, 537, 3, 537, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('tundra_high', 'temps', '15062023_05cm',
            temps_file, temps_worksheet, 'CW', 'EO', 3, 537, 3, 537, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
            # 15 cm
        standardization_method_1('palsa_low', 'temps', '15062023_15cm',
            temps_file, temps_worksheet, 'A', 'L', 3, 505, 3, 505, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('palsa_high', 'temps', '15062023_15cm',
            temps_file, temps_worksheet, 'A', 'AT', 3, 505, 3, 505, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('fen_low', 'temps', '15062023_15cm',
            temps_file, temps_worksheet, 'AY', 'BJ', 3, 444, 3, 444, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('fen_high', 'temps', '15062023_15cm',
            temps_file, temps_worksheet, 'AY', 'CR', 3, 444, 3, 444, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('tundra_low', 'temps', '15062023_15cm',
            temps_file, temps_worksheet, 'CW', 'DH', 3, 537, 3, 537, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('tundra_high', 'temps', '15062023_15cm',
            temps_file, temps_worksheet, 'CW', 'EP', 3, 537, 3, 537, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
            # 25 cm
        standardization_method_1('palsa_low', 'temps', '15062023_25cm',
            temps_file, temps_worksheet, 'A', 'M', 3, 306, 3, 306, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('palsa_high', 'temps', '15062023_25cm',
            temps_file, temps_worksheet, 'A', 'AU', 3, 306, 3, 306, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('fen_low', 'temps', '15062023_25cm',
            temps_file, temps_worksheet, 'AY', 'BK', 3, 444, 3, 444, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('fen_high', 'temps', '15062023_25cm',
            temps_file, temps_worksheet, 'AY', 'CS', 3, 444, 3, 444, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('tundra_low', 'temps', '15062023_25cm',
            temps_file, temps_worksheet, 'CW', 'DI', 3, 537, 3, 537, 'Date Time[]', 'Temp [C]', 0.01, 1.0)
        standardization_method_1('tundra_high', 'temps', '15062023_25cm',
            temps_file, temps_worksheet, 'CW', 'EQ', 3, 537, 3, 537, 'Date Time[]', 'Temp [C]', 0.01, 1.0)

    # Fluxes

    if params.std_fluxes:
            # June
        standardization_method_2('palsa_low', 'fluxes', '23062022', fluxes_file, fluxes_worksheet,
            'A', 'B', 'C', 2, 478, 2, 478, 2, 478, 'DateTime', 'CH4 Flux [mol/m2/day]', 0.01, 1.0)
        standardization_method_2('palsa_high', 'fluxes', '23062022_test', fluxes_file, fluxes_worksheet,
            'A', 'B', 'C', 2, 478, 2, 478, 2, 478, 'DateTime', 'CH4 Flux [mol/m2/day]', 0.01, 1.0)
        standardization_method_2('fen_low', 'fluxes', '23062022_test', fluxes_file, fluxes_worksheet,
            'A', 'B', 'C', 2, 478, 2, 478, 2, 478, 'DateTime', 'CH4 Flux [mol/m2/day]', 0.01, 1.0)
        standardization_method_2('fen_high', 'fluxes', '23062022_test', fluxes_file, fluxes_worksheet,
            'A', 'B', 'C', 2, 478, 2, 478, 2, 478, 'DateTime', 'CH4 Flux [mol/m2/day]', 0.01, 1.0)
        standardization_method_2('tundra_low', 'fluxes', '23062022_test', fluxes_file, fluxes_worksheet,
            'A', 'B', 'C', 2, 478, 2, 478, 2, 478, 'DateTime', 'CH4 Flux [mol/m2/day]', 0.01, 1.0)
        standardization_method_2('tundra_high', 'fluxes', '23062022_test', fluxes_file, fluxes_worksheet,
            'A', 'B', 'C', 2, 478, 2, 478, 2, 478, 'DateTime', 'CH4 Flux [mol/m2/day]', 0.01, 1.0)
            # August
        standardization_method_2('palsa_low', 'fluxes', '06082022', fluxes_file, fluxes_worksheet,
            'A', 'B', 'C', 2, 478, 2, 478, 2, 478, 'DateTime', 'CH4 Flux [mol/m2/day]', 0.01, 1.0)
        standardization_method_2('palsa_high', 'fluxes', '06082022_test', fluxes_file, fluxes_worksheet,
            'A', 'B', 'C', 2, 478, 2, 478, 2, 478, 'DateTime', 'CH4 Flux [mol/m2/day]', 0.01, 1.0)
        standardization_method_2('fen_low', 'fluxes', '06082022_test', fluxes_file, fluxes_worksheet,
            'A', 'B', 'C', 2, 478, 2, 478, 2, 478, 'DateTime', 'CH4 Flux [mol/m2/day]', 0.01, 1.0)
        standardization_method_2('fen_high', 'fluxes', '06082022_test', fluxes_file, fluxes_worksheet,
            'A', 'B', 'C', 2, 478, 2, 478, 2, 478, 'DateTime', 'CH4 Flux [mol/m2/day]', 0.01, 1.0)
        standardization_method_2('tundra_low', 'fluxes', '06082022_test', fluxes_file, fluxes_worksheet,
            'A', 'B', 'C', 2, 478, 2, 478, 2, 478, 'DateTime', 'CH4 Flux [mol/m2/day]', 0.01, 1.0)
        standardization_method_2('tundra_high', 'fluxes', '06082022_test', fluxes_file, fluxes_worksheet,
            'A', 'B', 'C', 2, 478, 2, 478, 2, 478, 'DateTime', 'CH4 Flux [mol/m2/day]', 0.01, 1.0)

    # Roots
    if params.std_roots:
        #
        standardization_method_3('palsa_low', 'roots', '15082023',
            roots_file, roots_worksheet, 'K', 'L', 23, 40, 23, 40, 'Date Time[]', 'Mass Density [kg/m^3]', 0.01, 1.0)
        standardization_method_3('palsa_high', 'roots', '15082023',
            roots_file, roots_worksheet, 'K', 'M', 23, 40, 23, 40, 'Date Time[]', 'Mass Density [kg/m^3]', 0.01, 1.0)
        standardization_method_3('fen_low', 'roots', '15082023',
            roots_file, roots_worksheet, 'K', 'L', 3, 20, 3, 20, 'Date Time[]', 'Mass Density [kg/m^3]', 0.01, 1.0)
        standardization_method_3('fen_high', 'roots', '15082023',
            roots_file, roots_worksheet, 'K', 'M', 3, 20, 3, 20, 'Date Time[]', 'Mass Density [kg/m^3]', 0.01, 1.0)
        standardization_method_3('tundra_low', 'roots', '15082023',
            roots_file, roots_worksheet, 'K', 'L', 43, 60, 43, 60, 'Date Time[]', 'Mass Density [kg/m^3]', 0.01, 1.0)
        standardization_method_3('tundra_high', 'roots', '15082023',
            roots_file, roots_worksheet, 'K', 'M', 43, 60, 43, 60, 'Date Time[]', 'Mass Density [kg/m^3]', 0.01, 1.0)

    return()


def standardize():

    read_site_data()

    return()

standardize()






