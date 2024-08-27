import os
import csv
import time
import data
import init
import params
import random
import shutil
import forcing
import cryogrid
import diffusion
import conversions
import numpy as np
from copy import copy
from os import listdir
from matplotlib import cm
from datetime import datetime
from datetime import timedelta
import matplotlib.pyplot as plt
from os.path import isfile, join
from openpyxl import load_workbook
from multiprocessing import Process, Queue
from mpl_toolkits.axes_grid1 import make_axes_locatable

# test 8/21/2024

def configure():

    cg_mode = 'configure'
    #

    # Configure dictionary
    configure_dictionary = dict([
        ('cg_mode', cg_mode), ('water_table_depth', params.water_table_depth),
        ('lateral', params.lateral), ('saturated', params.saturated),
        ('rain_fraction', params.rain_fraction), ('snow_fraction', params.snow_fraction),
        ('z0', params.z0), ('rootDepth', params.rootDepth), ('evaporationDepth', params.evaporationDepth),
        ('distance_reservoir', params.distance_reservoir), ('reservoir_contact_length', params.reservoir_contact_length)
    ])

    site_data_cg = data.read_site_data(params.site, params.cg_depths, params.cg_thicks)
    site_comp_cg = site_data_cg['site_comp']
    site_albedo_cg = site_data_cg['site_albedo']
    site_elevation_cg = site_data_cg['site_elevation']
    site_latitude_cg = site_data_cg['site_latitude']
    site_longitude_cg = site_data_cg['site_longitude']

    # Find the user-supplied composition profiles with the most completeness relative to the model grid
    data_keys = np.array(list(site_comp_cg.keys()))
    output_array = np.zeros_like(data_keys, dtype=bool)
    test_indices = np.char.find(data_keys, '_watm3m3')
    output_array[np.where(test_indices != -1)] = True
    watm3m3s = data_keys[output_array]
    completenesses = []
    for watm3m3 in watm3m3s:
        completenesses.append(len(set(site_comp_cg[watm3m3])))
    best_watm3m3 = watm3m3s[np.argmax(completenesses)]
    best_date = best_watm3m3.split('_')[-3]

    # Configure a moving (drying) water table, at a constant rate that fits the water table data
    water_table_keys = [s for s in list(site_comp_cg.keys()) if 'watertable' in s]
    dtimes, water_table_depths = [], []
    for wtk in water_table_keys:
        dtimes.append(conversions.ddmmyyyy2doy(wtk.split('_')[2]))
        water_table_depths.append(site_comp_cg[wtk])
    days_of_year, water_table_depths = np.array(conversions.sort_xylist_by_xlist(dtimes, water_table_depths))
    # Drying function
    m, b = np.polyfit(days_of_year, water_table_depths, 1)
    drying_begins, drying_rate = float(round((0.0 - b) / m)), float(round(m, 4))
    drying_begins, drying_rate = params.begin_drying, params.drying_rate
    # Read the file template
    f = open(params.cg_directory+'templates/WATER_FLUXES_LATERAL_template.m', 'r')
    lines = f.readlines()
    f.close()
    # Modify and replace the file
    line1 = '            drying_begins = ' + str(drying_begins) + ';\n'
    line2 = '                lateral.PARA.reservoir_elevation = ' + str(site_elevation_cg[list(site_elevation_cg.keys())[0]] - 0.05) + ' - days_since_drying * ' + str(drying_rate) + ';\n'
    line3 = '                lateral.PARA.reservoir_elevation = ' + str(site_elevation_cg[list(site_elevation_cg.keys())[0]] - 0.05) + ';\n'
    line4 = '                lateral.PARA.reservoir_elevation = ' + str(site_elevation_cg[list(site_elevation_cg.keys())[0]] - 0.05) + ' - (300.0 - drying_begins) * ' + str(drying_rate) + ';\n'
    lines[207], lines[213], lines[217], lines[220] = line1, line2, line3, line4
    new_WATER_FLUXES_LATERAL = params.cg_directory + 'CryoGridCommunity_source/source/TIER_1_processes/WATER_FLUXES_LATERAL.m'
    f = open(new_WATER_FLUXES_LATERAL, 'w')
    f.writelines(lines)
    f.close()

    # Calculate a field capacity = average water content of sub-saturated layers
    field_capacity = np.mean(
        site_comp_cg[params.site + '_' + best_date + '_watm3m3_VvZ']
        [site_comp_cg[params.site + '_' + best_date + '_watm3m3_VvZ'] <= 0.5])
    print(field_capacity)
    field_capacity = 0.35

    filename = params.cg_directory + 'templates/template.xlsx'
    # Make a site directory if it doesn't already exist
    if not os.path.exists(params.cg_directory + 'CryoGridCommunity_results/' + params.site + '/' + params.site + '_' + params.cryogrid_name):
        os.mkdir(params.cg_directory + 'CryoGridCommunity_results/' + params.site + '/' + params.site + '_' + params.cryogrid_name)

    filename2 = params.cg_directory + 'CryoGridCommunity_results/' + params.site + '/' + params.site + '_' + params.cryogrid_name + '/' + params.site + '_' + params.cryogrid_name + '.xlsx'
    shutil.copy(filename, filename2)

    wb = load_workbook(filename2)
    ws = wb.active

    # geography
    ws['B13'] = float(site_latitude_cg[list(site_latitude_cg.keys())[0]])
    ws['B14'] = float(site_longitude_cg[list(site_longitude_cg.keys())[0]])
    ws['B15'] = float(site_elevation_cg[list(site_elevation_cg.keys())[0]])

    # meteorology file
    ws['B47'] = 'Churchill_' + params.site + '_T0_ERA5_Land_1950_2022.mat'

    # albedo
    ws['B141'] = float(site_albedo_cg[list(site_albedo_cg.keys())[0]])
    ws['B166'] = float(site_albedo_cg[list(site_albedo_cg.keys())[0]])

    # elevation
    ws['B211'] = float(site_elevation_cg[list(site_elevation_cg.keys())[0]])
    ws['B75'] = float(site_elevation_cg[list(site_elevation_cg.keys())[0]]) + 1.5
    ws['B76'] = float(site_elevation_cg[list(site_elevation_cg.keys())[0]]) + 1.5 - 6.5

    # extra stratigraphy info
    ws['I117'] = 'air'
    ws['J117'] = 'sat %'

    # save_date
    ws['B63'] = params.save_date
    ws['B79'] = params.save_date

    # start_time
    ws['C49'] = int((params.start_time)[0])
    ws['D49'] = int((params.start_time)[1])
    ws['E49'] = int((params.start_time)[2])

    # end_time
    ws['C50'] = int((params.end_time)[0])
    ws['D50'] = int((params.end_time)[1])
    ws['E50'] = int((params.end_time)[2])

    # rain and snow fractions
    ws['B51'] = params.rain_fraction
    ws['B52'] = params.snow_fraction

    # plants
    ws['B143'] = params.z0
    ws['B144'] = params.rootDepth
    ws['B145'] = params.evaporationDepth

    # lateral reservoir
    # Lateral flow is ON by default. Turn this OFF if params.lateral = False.
    if params.lateral == False:
        ws['C39'] = 'END'
        ws['C40'] = 'END'
        ws['D39'] = ''
        ws['D40'] = ''
    # lateral reservoir elevation [m]
    ws['B223'] = float(site_elevation_cg[params.site + '_' + best_date + '_elevation']) - params.water_table_depth
    # lateral reservoir distance [m]
    ws['B226'] = params.distance_reservoir
    # lateral reservoir contact length [m]
    ws['B227'] = params.reservoir_contact_length

    # stratigraphy
    ws.insert_rows(118, len(params.cg_depths))
    for i in range(0, len(params.cg_depths)):

        ws['B' + str(118 + i)] = params.cg_depths[i]
        ws['C' + str(118 + i)] = float("{:.3f}".format(site_comp_cg[params.site + '_' + best_date + '_watm3m3_VvZ'][i]))
        if params.saturated:
            ws['C' + str(118 + i)] = 1.0 - \
                                     (float(
                                         "{:.3f}".format(site_comp_cg[params.site + '_' + best_date + '_minm3m3_VvZ'][i]))
                                      + float("{:.3f}".format(
                                                 site_comp_cg[params.site + '_' + best_date + '_socm3m3_VvZ'][i])))
        ws['D' + str(118 + i)] = float("{:.3f}".format(site_comp_cg[params.site + '_' + best_date + '_minm3m3_VvZ'][i]))
        ws['E' + str(118 + i)] = float("{:.3f}".format(site_comp_cg[params.site + '_' + best_date + '_socm3m3_VvZ'][i]))
        ws['F' + str(118 + i)] = int(1)
        ws['G' + str(118 + i)] = float("{:.3f}".format(field_capacity))
        ws['I' + str(118 + i)] = float("{:.3f}".format(site_comp_cg[params.site + '_' + best_date + '_airm3m3_VvZ'][i]))
        ws['J' + str(118 + i)] = 100.0 * float(
            "{:.3f}".format(site_comp_cg[params.site + '_' + best_date + '_airm3m3_VvZ'][i] +
                            site_comp_cg[params.site + '_' + best_date + '_watm3m3_VvZ'][i]))

    wb.save(filename=filename2)

    f = open(params.cg_directory + 'templates/run_CG_template.m', 'r')
    lines = f.readlines()
    lines[7] = "result_path = '../CryoGridCommunity_results/" + params.site + "/';"
    n = 10
    lines[
        n - 1] = "run_name = '" + params.site + '_' + params.cryogrid_name + "'; %run_name = name of parameter file (without file extension) and name of subfolder (in result_path) within which it is located  # n is the line number you want to edit; subtract 1 as indexing of list starts from 0"
    f.close()

    new_run_CG = params.cg_directory + 'CryoGridCommunity_run/run_CG.m'
    f = open(new_run_CG, 'w')
    f.writelines(lines)
    f.close()

    shutil.copyfile(params.cg_directory + 'templates/CONSTANTS_excel_template.xlsx',
                    params.cg_directory + 'CryoGridCommunity_results/' + params.site + '/' + params.site + '_' + params.cryogrid_name + '/CONSTANTS_excel.xlsx')

    return(configure_dictionary)

def run():

    print('Running CryoGrid')
    os.system('matlab -nosplash -nodesktop -batch "script1"')

    return()


def process():
    print('Processing CryoGrid output')

    ### script2.m loads the desired output file, and makes + writes clean output files into the 'temp' folder

    # Prepare a 'temp' folder to temporarily store the processed output files
    if not os.path.exists(params.main_directory + 'CryoGrid/CryoGridCommunity_results/temp'):
        os.mkdir(params.main_directory + 'CryoGrid/CryoGridCommunity_results/temp')

    # Re-write script2.m to include the site name in the 'load' command
    script2_data = np.array(
        read_csv_array(r'C:\Users\Jeremy\Desktop\Churchill_Data\CryoGrid\templates\script2_template.m'))
    new_lines = []
    for d in range(0, len(script2_data)):
        line = script2_data[d][0]
        if 'load' in line:
            test1, test2 = line.find('_results'), line.find('_drained')
            half1, half2, add = line[0:test1 + 8], line[test2:], '\\' + params.site + '\\' + params.site + '_' + params.cryogrid_name + '\\' + params.site + '_' + params.cryogrid_name
            line = half1 + add + half2
        new_lines.append(line)

    with open(r'C:\Users\Jeremy\PycharmProjects\microbeCH4\script2.m', "w") as current_file:
        for line in new_lines:
            current_file.write(line + '\n')

    # Run script2.m
    os.system('matlab -nosplash -nodesktop -wait -batch "script2"')
    print(' ')
    print('Loading and processing CryoGrid output (script2.mat): ')
    print(new_lines)

    # Prepare a site-specific directory for the clean output files,
    # and copy clean output files to the site-specific directory
    print('Making processed CryoGrid output: ', params.main_directory + 'CryoGrid/' + params.site + '/clean_output')
    if not os.path.exists(params.main_directory + 'CryoGrid/CryoGridCommunity_results/' + params.site + '/' + params.site + '_' + params.cryogrid_name + '/clean_output'):
        os.mkdir(params.main_directory + 'CryoGrid/CryoGridCommunity_results/' + params.site + '/' + params.site + '_' + params.cryogrid_name + '/clean_output')
    recursive_copy(params.main_directory + 'CryoGrid/CryoGridCommunity_results/temp',
                   params.main_directory + 'CryoGrid/CryoGridCommunity_results/' + params.site + '/' + params.site + '_' + params.cryogrid_name + '/clean_output')

    return()


def read_csv_array(filename):

    with open(filename, newline='') as csvfile:
        data = list(csv.reader(csvfile))

    return(data)


def recursive_copy(src, dest):

    for item in os.listdir(src):
        file_path = os.path.join(src, item)

        # if item is a file, copy it
        if os.path.isfile(file_path):
            shutil.copy(file_path, dest)

        # else if item is a folder, recurse
        elif os.path.isdir(file_path):
            new_dest = os.path.join(dest, item)
            os.mkdir(new_dest)
            recursive_copy(file_path, new_dest)


def cggrid2modelgrid(cg_forcing):

    # New CryoGrid forcing dictionary, with profiles interpolated to [n_model_layers x n_CryoGrid_timesteps]
    cg_forcing2 = cg_forcing.copy()

    # n_CryoGrid_timesteps
    first_key = list(cg_forcing.keys())[0]
    ncgt = np.shape(cg_forcing[first_key])[1]

    # n_model_layers
    layer_thicknesses, depths = init.define_layers()
    depths=-1.0*depths

    # Interpolate the data to the model depth grid
    for key in cg_forcing:
        key_dimensions = len(np.shape(cg_forcing[key]))

        ir = 14 # Number of CryoGrid layers to utilize for interpolation (from the surface)

        if key_dimensions == 1: # Only interpolate time or space profiles, not e.g. depth_data or time_data
            data_interp = cg_forcing[key]

        if key_dimensions == 2: # Only interpolate time x space profiles, not e.g. depth_data or time_data
            data_interp = np.zeros((len(depths), ncgt))
            for i in range(0, np.shape(data_interp)[1]):
                xnew, data_interp[:, i] = forcing.interp2grid(cg_forcing['depth_data'][0:ir], cg_forcing[key][0:ir, i], depths)

        if key_dimensions == 3:  # Only interpolate type x time x space profiles, not e.g. depth_data or time_data
            data_interp = np.zeros( (np.shape(cg_forcing[key])[0] , len(depths) , ncgt))
            for k in range(0, np.shape(data_interp)[0]):
                for i in range(0, np.shape(data_interp)[2]):
                    xnew, data_interp[k, :, i] = forcing.interp2grid(cg_forcing['depth_data'][0:ir], cg_forcing[key][k, 0:ir, i], depths)

        cg_forcing2[key] = data_interp

    # The model depths are the new depths
    cg_forcing2['depth_data'] = depths

    return(cg_forcing2)


def profile_plots(axis,data,tempdata,satdata,levels,xlabel,timedata,depthdata,color_scheme,fig,temp_nan,secondary):

    axis.set_xlim([150, 320])

    X, Y = np.meshgrid(timedata, depthdata)
    data2 = copy(data)
    ###data2[tempdata < 0.0] = np.nan
    #c = axis.contourf(X, Y, data2, levels=levels, cmap=color_scheme)
    temp_mask = 1.0 + 0.0 * tempdata
    temp_mask[tempdata > 0.0] = np.nan
    c = axis.contourf(X, Y, data2, cmap=color_scheme, levels=100)
    axis.set_ylim([-0.60, 0.0])
    axis.set_xlabel('Day of Year',fontsize=24)
    divider = make_axes_locatable(axis)
    if secondary == False:
        cax = divider.append_axes('bottom', size='5%', pad=1.2)
        cbar = fig.colorbar(c, cax=cax, orientation='horizontal')
        cbar.ax.tick_params(axis="both", labelsize=18, rotation=45)
        cbar.set_label(xlabel, rotation=0, fontsize=22, fontweight="bold")
    if secondary:
        cax = divider.append_axes('bottom', size='5%', pad=-2.3)
        fig.add_axes(cax)
        cbar = fig.colorbar(c, cax=cax, orientation='horizontal')
        cbar.ax.tick_params(axis="both", labelsize=18, rotation=45)
        cbar.set_label(xlabel, rotation=0, fontsize=22, fontweight="bold")
        cbar.set_ticks([])
    axis.xaxis.set_tick_params(labelsize=20)
    axis.yaxis.set_tick_params(labelsize=20)
    axis.set_ylabel('Depth (m)',fontsize=22)

    axis.contourf(X, Y, temp_mask, colors = 'white', hatches = '/', alpha = 0.5)
    if temp_nan:
        axis.contourf(X, Y, temp_mask, colors='white')
    c = axis.contour(X, Y, tempdata, linewidths=2.0, levels=[-400,0.0],colors='black')

    c = axis.contour(X, Y, satdata, linewidths=4.0, levels=[0.0, 99.0],colors='white',alpha=0.5)
    c = axis.contour(X, Y, satdata, linewidths=2.0, levels=[0.0, 99.0],colors='blue',alpha=0.5,linestyles='--')

    return()

#configure()

#run()

#process()

forcing.all_forcing()