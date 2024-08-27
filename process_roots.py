from skimage.morphology import skeletonize, thin, disk, erosion, dilation, remove_small_objects
from scipy.interpolate import CubicSpline
from scipy.signal import argrelextrema
from openpyxl import load_workbook
from colorthief import ColorThief
import matplotlib.pyplot as plt
from skimage.util import invert
import matplotlib.pyplot as plt
from numpy import random
from skimage import data
import skimage.filters
from PIL import Image
import imageio as iio
import skimage.color
import numpy as np
import glob
import cv2
import csv
import os


def format_axis1(ax):

    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    return()


def format_axis2(ax):

    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.set_xticks([])
    ax.set_yticks([])

    return()


def find_nearest_white(img, target):

    nonzero = cv2.findNonZero(img)
    distances = np.sqrt((nonzero[:,:,0] - target[0]) ** 2 + (nonzero[:,:,1] - target[1]) ** 2)
    nearest_index = np.argmin(distances)
    return nonzero[nearest_index]


def read_csv_header(filename, column_idx, var_type, header_lines):
    with open(filename) as f:
        reader = csv.reader(f)
        if header_lines != 0:
            for h in range(0,header_lines):
                header = next(reader)
        vals = []
        for row in reader:
            if var_type == 'string':
                val = row[column_idx]
            if var_type == 'integer':
                val = int(row[column_idx])
            if var_type == 'float':
                if row[column_idx] == '':
                    val = -9999.0
                else:
                    val = float(row[column_idx])
            vals.append(val)
    return vals

def write2excel(sample, column_offset, value):

    # test

    plot_str = sample.split('_')[0]
    site_str, plot_num = plot_str[0], int(plot_str[1:]) - 1

    depths = np.array(
        [2.5, 5.0, 7.5, 10.0, 12.5, 15.0, 17.5, 20.0, 22.5, 25.0, 27.5, 30.0, 32.5, 35.0, 37.5, 40.0, 42.5])
    depth_str = sample.split('_')[1]
    depth = float(depth_str.split('p')[0]) + float(depth_str.split('p')[1]) / 10.0
    if depth in depths:

        depth_idx = np.where(depths == depth)[0][0]

        f_row, p_row, t_row = 0, 18, 36
        if site_str == 'F': row = str(f_row + depth_idx + 0)
        if site_str == 'P': row = str(p_row + depth_idx + 1)
        if site_str == 'T': row = str(t_row + depth_idx + 2)

        letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                   'V', 'W', 'X', 'Y', 'Z']
        volume_idx = [0, 30, 45, 60, 75, 90, 105, 120, 135, 15]

        column = volume_idx[plot_num] + column_offset
        print(column)
        if column <= 26: column_letter = letters[column - 1]
        if column > 26: column_letter = letters[int(np.floor(column / 26) - 1)] + letters[column % 26]
        print(column, letters[int(np.floor(column / 26) - 1)], letters[column % 26], column_letter)

        #filename2 = 'C:/Users/Jeremy/Desktop/Churchill_Data/siteData/raw/roots_master.xlsx'
        #wb = load_workbook(filename2)
        #ws = wb.active

        #ws[column_letter + row] = value
        #print(row, column_letter, ' -> ', value)
        #print(' ')
        #wb.save(filename=filename2)

    return()

#ax4.text(0.15, 1.00, 'Root:Peat \n' + str(format(root2peat_pixel_ratio, ".4f")), ha='center', va='center', fontsize=10, weight='bold', color='black')
#ax4.text(0.45, 1.00, 'Tortuosity \n' + str(format(mean_tortuosity, ".4f")), ha='center', va='center', fontsize=10, weight='bold', color='black')
#ax4.text(0.75, 1.00, 'X-Section \n' + str(format(total_xsection, ".4f")), ha='center', va='center', fontsize=10, weight='bold', color='black')
#ax4.text(0.15, -1.00, 'Length \n' + str(format(total_root_length, ".4f")), ha='center', va='center', fontsize=10, weight='bold', color='black')
#ax4.text(0.45, -1.00, 'Color \n' + str(colors[2][1]), ha='center', va='center', fontsize=10, weight='bold', color='black')

every_nth = 100 # every nth centroid
n_tortuosities = 10000

img_path = 'C:/Users/Jeremy\Desktop/Churchill_Data/siteData/raw/roots_bg_removed/T/2/T2_12p5.jpg'

site = 'P'

plot_nums = os.listdir('C:/Users/Jeremy\Desktop/raw_bu/roots_bg_removed/' + site)
plot_nums = ['9']

for p in plot_nums:
    plot_depths = os.listdir('C:/Users/Jeremy\Desktop/raw_bu/roots_bg_removed/' + site + '/' + p)
    for img in plot_depths:

        # Clear all variables (to prevent writing repeat values)
        root2peat_pixel_ratio, mean_tortuosity, mean_sizes_low_end, mean_sizes_all, total_root_pixel_length_low_end, \
        total_root_pixel_length, plot_str = 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 'None'

        img_path = 'C:/Users/Jeremy\Desktop/raw_bu/roots_bg_removed/' + site + '/' + p + '/' + img
        img_path = 'C:/Users/Jeremy/Desktop/raw_bu/roots_bg_removed/T/2/T2_10p0.JPG'
        print('plot depths: ', plot_depths)
        print('image: ', img_path)

        plot_str = (img_path.split('/')[-1]).split('.')[0]

        # Load image and extract R, G, B layers
        img = cv2.imread(img_path)
        img = cv2.GaussianBlur(img, (0,0), sigmaX=10, sigmaY=10, borderType = cv2.BORDER_DEFAULT)
        img_b, img_g, img_r = img[:, :, 0], img[:, :, 1], img[:, :, 2]
        img = Image.open(img_path)

        # Reduce the image to 3 colors (background, peat, roots)
        result = img.convert('P', palette=Image.ADAPTIVE, colors=3)
        colors = result.convert('RGB').getcolors()
        result = np.asarray(result.convert('RGB')).astype('float32')
        print(result.shape)

        # Make the root mask
        print('Making root mask')
        cidx = 2
        test = np.logical_and((result[:,:,0] == colors[cidx][1][0]), np.logical_and((result[:,:,1] == colors[cidx][1][1]), (result[:,:,2] == colors[cidx][1][2])))
        root_mask_0 = 0 * np.copy(img_r)
        root_mask_0[test] = 1
        # Cleanup the root mask
        root_mask_1 = root_mask_0
        #root_mask_1 = remove_small_objects(root_mask_1, 1000)

        # Make the peat mask
        print('Making peat mask')
        cidx = 1
        test = np.logical_and((result[:,:,0] == colors[cidx][1][0]), np.logical_and((result[:,:,1] == colors[cidx][1][1]), (result[:,:,2] == colors[cidx][1][2])))
        peat_mask_0 = 0 * np.copy(img_r)
        peat_mask_0[test] = 1
        print('Cleaning peat mask')
        # Cleanup the peat mask
        peat_mask_1 = peat_mask_0
        #peat_mask_1 = dilation(peat_mask_0, disk(3.0))
        peat_mask_1 = erosion(peat_mask_0, disk(2.0))
        #peat_mask_1 = remove_small_objects(peat_mask_1, 1000)

        # Make the background mask
        print('Making background mask')
        cidx = 0
        test = np.logical_and((result[:,:,0] == colors[cidx][1][0]), np.logical_and((result[:,:,1] == colors[cidx][1][1]), (result[:,:,2] == colors[cidx][1][2])))
        bg_mask_0 = 0 * np.copy(img_r)
        bg_mask_0[test] = 1
        print('Cleaning background mask')
        # Cleanup
        #bg_mask_1 = bg_mask_0
        bg_mask_1 = dilation(bg_mask_0, disk(3.0))
        #peat_mask_1 = erosion(peat_mask_0, disk(1.0))
        #peat_mask_1 = remove_small_objects(peat_mask_1, 1000)

        # Make sure that no root pixels coincide with background pixels
        root_mask_1[bg_mask_1 == 1] = 0
        root_mask_1 = dilation(root_mask_0, disk(1.0))
        #root_mask_1 = erosion(root_mask_1, disk(1.0))

        # Root : Peat pixel ratio
        root2peat_pixel_ratio = np.sum(root_mask_1) / np.sum(peat_mask_1)
        print('Root:Peat ratio: ', root2peat_pixel_ratio)

        ## Root midlines
        # Make a mask of the root pixel midlines, for size and tortuosity determination
        root_skele_0 = skeletonize(root_mask_0)
        # Columns and rows spanned by the midline pixels
        root_skele_0_rows, root_skele_0_cols = np.where(root_skele_0 == 1.0)[0], np.where(root_skele_0 == 1.0)[1]
        # Make sure that no midline pixels coincide with peat pixels
        root_skele_0[peat_mask_0 == 1] = 0

        # Calculate the total root length (sum of midline pixels)
        total_root_pixel_length = np.sum(root_skele_0)

        # Geometric center coordinates and pixel extent of the root pixels, for cropped plot ranges
        row_mid = np.mean(root_skele_0_rows)
        col_mid = np.mean(root_skele_0_cols)
        row_span = (max(root_skele_0_rows) - min(root_skele_0_rows)) / 2.0
        col_span = (max(root_skele_0_cols) - min(root_skele_0_cols)) / 2.0
        max_span = min([row_span, col_span])
        max_span = 700

        # Initialize plot grid
        fig = plt.figure(figsize=(10.0, 8.0))
        gsc = fig.add_gridspec(2, 2, width_ratios=np.ones(2), height_ratios=[2, 1])

        # Plot showing root/peat/background differention, size measurement points, and tortuosity trajectories
        ax3 = fig.add_subplot(gsc[0, 1])
        ax3.set_xlim([col_mid - max_span, col_mid + max_span])
        ax3.set_ylim([row_mid - max_span, row_mid + max_span])
        format_axis2(ax3)
        ax3.imshow(img)

        # Plot showing root size frequency distribution, with fitted polynomial for extrema identification
        ax = fig.add_subplot(gsc[0, 0])
        ax.set_xlim([col_mid - max_span, col_mid + max_span])
        ax.set_ylim([row_mid - max_span, row_mid + max_span])
        format_axis2(ax)
        ax.imshow(img)

        # Plot overlay-able peat, root, root skeleton, and background layers
        cmap = plt.get_cmap('Reds')
        peat_transparent = (cmap(peat_mask_1 / np.max(peat_mask_1)) * 255).astype(np.uint8)
        peat_transparent[:, :, 3] = 255 + 0.0 * (peat_mask_1 > 0) * 255
        #
        cmap = plt.get_cmap('Greens')
        root_transparent = (cmap(root_mask_1 / np.max(root_mask_1)) * 255).astype(np.uint8)
        root_transparent[:, :, 3] = 255 + 0.0 * (root_mask_1 > 0) * 255
        #
        cmap = plt.get_cmap('Greys')
        root_skele_transparent = (cmap(root_skele_0 / np.max(root_skele_0)) * 255).astype(np.uint8)
        root_skele_transparent[:, :, 3] = (root_skele_0 > 0) * 255
        #
        cmap = plt.get_cmap('Greys')
        bg_transparent = (cmap(bg_mask_1 / np.max(bg_mask_1)) * 255).astype(np.uint8)
        bg_transparent[:, :, 3] = (bg_mask_1 > 0) * 255
        #
        ax3.imshow(peat_transparent, alpha=1.0)
        ax3.imshow(root_transparent, alpha=1.0)
        #plt.imshow(bg_transparent, alpha = 1.0)
        #ax3.imshow(img, alpha = 0.5)
        ax3.imshow(root_skele_transparent, alpha = 1.0)

        #binary_stack = root_mask_1
        reverse_image = np.copy(root_mask_1)
        reverse_image[root_mask_1 > 0] = 0
        reverse_image[root_mask_1 == 0] = 1
        #binary_stack[root_mask_1 > 0.0] = 1.0
        median_points = cv2.findNonZero(root_skele_0.astype(float))
        widths = []
        print('Processing...')
        count = 0
        for i in range(0,len(median_points), every_nth):
            progress = 100.0*i/len(median_points)
            if round(progress, 1) % 10.0 == 0:
                print(progress)
            median_point = median_points[i][0]
            target = (median_point[0],median_point[1])
            coords = find_nearest_white(reverse_image, target)[0]
            #ax3.plot([target[0],coords[0]],[target[1],coords[1]],linestyle='-', linewidth = 5.0, color = 'black')
            #ax3.plot([target[0], coords[0]], [target[1], coords[1]], linestyle='-', linewidth = 3.0, color='white')
            if count%10 == 0:
                ax3.plot([coords[0]], [coords[1]], marker = '.', color = 'purple', markersize = 10.0)
                ax3.plot([coords[0]], [coords[1]], marker='.', color='white', markersize=5.0)
            width = np.sqrt(abs(coords[0] - target[0])**2.0 + abs(coords[1] - target[1])**2.0)
            widths.append(width)
            count = count + 1

        ratios = [] # List of tortuosity values
        for t in range(0, n_tortuosities): # Number of tortuosity measurements to make

            # Choose a random point on the root midline mask
            rand_idx = random.randint(len(root_skele_0_rows))
            rand_row, rand_col = root_skele_0_rows[rand_idx], root_skele_0_cols[rand_idx]

            rand_rows, rand_cols = [], [] # Keep a record of the row and column coordinates encountered by this walk
            coverage = 0 * root_skele_0 # Keep a record of row and column coordinates encountered by all walks
            deadended, looped = False, False # Reset the dead-end and looping checks
            # Walk along the trajectory until dead-ending or looping occurs
            i = 0
            while (deadended == False) and (looped == False) and \
                rand_row + 1 < root_skele_0.shape[0] and rand_col + 1 < root_skele_0.shape[1]:

                # Clockwise search for midline pixels around the current midline point, starting at 12-o-clock position
                p0, p1, p2, p3, p4, p5, p6, p7 = \
                    root_skele_0[rand_row - 1, rand_col], root_skele_0[rand_row - 1, rand_col + 1], \
                    root_skele_0[rand_row, rand_col + 1], root_skele_0[rand_row + 1, rand_col + 1], \
                    root_skele_0[rand_row + 1, rand_col], root_skele_0[rand_row + 1, rand_col - 1], \
                    root_skele_0[rand_row, rand_col - 1], root_skele_0[rand_row - 1, rand_col - 1]

                # Mark the current position on the coverage map
                coverage[rand_row, rand_col] = 1

                # Shift the position to the first found adjacent midline pixel
                moved = 0
                if p0 and (coverage[rand_row - 1, rand_col] == 0) and moved == 0:
                    rand_row = rand_row - 1
                    moved = 1
                if p1 and (coverage[rand_row - 1, rand_col + 1] == 0) and moved == 0:
                    rand_row = rand_row - 1
                    rand_col = rand_col + 1
                    moved = 1
                if p2 and (coverage[rand_row, rand_col + 1] == 0) and moved == 0:
                    rand_col = rand_col + 1
                    moved = 1
                if p3 and (coverage[rand_row + 1, rand_col + 1] == 0) and moved == 0:
                    rand_row = rand_row + 1
                    rand_col = rand_col + 1
                    moved = 1
                if p4 and (coverage[rand_row + 1, rand_col] == 0) and moved == 0:
                    rand_row = rand_row + 1
                    moved = 1
                if p5 and (coverage[rand_row + 1, rand_col - 1] == 0) and moved == 0:
                    rand_row = rand_row + 1
                    rand_col = rand_col - 1
                    moved = 1
                if p6 and (coverage[rand_row, rand_col - 1] == 0) and moved == 0:
                    rand_col = rand_col - 1
                    moved = 1
                if p7 and (coverage[rand_row - 1, rand_col - 1] == 0) and moved == 0:
                    rand_row = rand_row - 1
                    rand_col = rand_col - 1
                    moved = 1
                # Record the row and column coordinates of the new position
                rand_rows.append(rand_row)
                rand_cols.append(rand_col)

                # Check whether the walk has reached a dead-end or has looped back onto itself
                deadended = (i >= 3) and (rand_rows[-1] == rand_rows[-3]) and (rand_cols[-1] == rand_cols[-3])
                looped = (i >= 2) and (rand_rows[-1] == rand_rows[0]) and (rand_cols[-1] == rand_cols[0])

                # Make a tortuosity measurement if the walk has advanced by at least two pixels
                if i >= 2:
                    indirect = len(rand_rows)
                    direct = np.sqrt((rand_rows[-1] - rand_rows[0])**2. + (rand_cols[-1] - rand_cols[0])**2.)
                    ratio = indirect / direct

                i += 1

            # Record the final tortuosity measurement if the walk > 50 pixels, didn't loop, and has reasonable values
            if indirect > 50 and looped == False and (1.0 < ratio < 5.0):
                ratios.append(ratio)
                # Dynamically plot the walk trajectory
                ax3.plot(rand_cols, rand_rows, linestyle = '-', linewidth = 2.0, alpha = 0.5, color = 'red')

        # Calculate the mean tortuosity value
        mean_tortuosity = np.mean(ratios)
        print('Mean tortuosity: ', mean_tortuosity)

        widths = np.array(widths)
        mean_sizes_all = np.mean(widths)
        binwidth = 0.5
        histo = np.histogram(widths, bins=10)
        size_counts = histo[0]
        size_sizes = histo[1][0:-1]
        total_counts = np.sum(size_counts) # Number of radii measured
        x, y = histo[1][1:], histo[0]
        xs, cs = np.arange(0.0, 10.0, 0.1), CubicSpline(x, y)

        ax2 = fig.add_subplot(gsc[1, :])
        ax2.set_ylim([0, 1.2 * max(histo[0])])
        ax2.set_xlim([min(test[1]), max(histo[1])])
        format_axis1(ax2)
        ax2.set_xlabel('Root Diameter (pixels)')
        ax2.set_ylabel('Counts (N)')

        #test = ax2.plot(x - (x[1] - x[0])/2.0, y, color = 'blue')

        maxima_idx = argrelextrema(cs(xs), np.greater)[0]
        maxima = (xs - (x[1] - x[0])/2.0)[maxima_idx]

        minima_idx = argrelextrema(cs(xs), np.less)[0]
        minima = (xs - (x[1] - x[0])/2.0)[minima_idx]

        if len(minima_idx) > 0:

            print(size_counts)
            print(size_sizes)
            print('minima_idx: ', minima_idx)
            print('minima: ', minima)
            print(minima[0])
            print(size_counts[size_sizes < minima[0]])

            ax2.plot(xs - (x[1] - x[0])/2.0, cs(xs), linestyle = '-', linewidth = 3.0, color='grey')
            low_end_sizes_idx = (x - (x[1] - x[0])/2.0) < minima[0]
            low_end_sizes = (x - (x[1] - x[0])/2.0)[(x - (x[1] - x[0])/2.0) < minima[0]] # List of measured root ending radii
            print(low_end_sizes)
            low_end_counts = y[low_end_sizes_idx]
            print(low_end_counts)

            #ax2.plot(low_end_sizes, low_end_counts, linestyle = '-', linewidth = 3.0, color='red')
            #ax2.fill_between(low_end_sizes, 0.0 * low_end_counts, low_end_counts, alpha=0.2, color='red')
            ax2.set_ylim([0, 0.5 * max(cs(xs))])

            print('total counts: ', total_counts)
            print('low end sizes: ', low_end_sizes)

            mean_sizes_low_end = np.average(low_end_sizes, weights = low_end_counts)

            print('mean_sizes_low_end: ', mean_sizes_low_end)
            print('mean_sizes_all: ', mean_sizes_all)

            # Proportion of measured radii falling in each low-end size bin to all radii measured. Assumed to be the proportion of
            # measured radii falling in each low-end size bin to all radii along the total root length, due to random sampling
            low_end_frequencies = low_end_counts / total_counts
            print(low_end_frequencies)

            total_root_pixel_length_low_end = total_root_pixel_length * low_end_frequencies[0]

            print('total_root_pixel_length_low_end: ', total_root_pixel_length_low_end)
            print('total_root_pixel_length: ', total_root_pixel_length)

            print('low end sizes / total counts: ', low_end_counts / total_counts)
            total_pixel_xsection = np.pi * (np.sum(total_root_pixel_length * low_end_frequencies))**2.0 # Total root ending x-section [pixels^2]
            print(total_pixel_xsection)
            #ax2.set_aspect(0.05)
            #for t in range(0, len(maxima)):
            #    ax2.plot([xs[maxima[t]] - (x[1] - x[0])/2.0], [cs(xs)[maxima[t]]], marker = 'x', markersize = 5.0)
            #    ax2.text(xs[maxima[t]] - (x[1] - x[0])/2.0, cs(xs)[maxima[t]], str(round(xs[test[t]] - (x[1] - x[0])/2.0, 3)),
            #             ha='center', va='center', fontsize=12, weight='bold', color='black',
            #             bbox=dict(facecolor='white', alpha=1.0, boxstyle='round', edgecolor='orange'))

        ax4 = fig.add_subplot(gsc[1, 1])
        ax4.plot([0.0], [0.0])
        ax4.set_ylim([-2.0, 2.0])
        ax4.set_xlim([0.0, 1.0])
        ax4.spines['top'].set_visible(False)
        ax4.spines['bottom'].set_visible(False)
        ax4.set_axis_off()

        ax.set_xlabel('Root:Peat \n' + str(format(root2peat_pixel_ratio, ".4f")), weight='bold', color='black')
        ax3.set_xlabel('Tortuosity \n' + str(format(mean_tortuosity, ".4f")), weight='bold', color='red')
        ax3.set_ylabel('Total Length \n' + str(format(total_root_pixel_length, ".4f")), weight='bold', color='green')
        ax2.text(mean_sizes_low_end, 0.25 * max(cs(xs)), 'Low Mean Diam \n' + str(format(mean_sizes_low_end, ".4f")), weight='bold', color='purple')
        ax2.text(mean_sizes_all, 0.40 * max(cs(xs)), 'All Mean Diam\n' + str(format(mean_sizes_all, ".4f")), weight='bold', color='purple')

        # Write calculated quantities to the excel file at the specified column offset
        write2excel(plot_str, 7, str(format(root2peat_pixel_ratio, ".4f"))) # Root:Peat area ratio
        write2excel(plot_str, 8, str(format(mean_tortuosity, ".4f"))) # Mean tortuosity
        write2excel(plot_str, 9, str(format(mean_sizes_low_end, ".4f"))) # Mean pixel diameter of root ending measurements
        write2excel(plot_str, 10, str(format(mean_sizes_all, ".4f"))) # Mean pixel diameter of all measurements
        write2excel(plot_str, 11, str(format(total_root_pixel_length_low_end, ".4f"))) # Total pixel length of root endings
        write2excel(plot_str, 12, str(format(total_root_pixel_length, ".4f"))) # Total pixel length of all roots
        write2excel(plot_str, 13, str(colors[2][1])) # Mean root color (RGB)

        # colors[2][1]
        # root2peat_ratio
        # mean_tortuosity
        # total_xsection
        # total_root_length

        #print(colors[2][1])

        #ax4.plot([0.5], [0.0], marker = 'o', color = (colors[2][1][0]/255.0, colors[2][1][1]/255.0, colors[2][1][2]/255.0))

        plt.savefig('C:/Users/Jeremy/Desktop/Churchill_Data/siteData/raw/root_test.png',bbox_inches='tight')
        plt.show()

        stop